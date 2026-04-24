from __future__ import annotations

from collections import OrderedDict
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


INPUT_PATH = Path(r"C:\Users\60893\Downloads\BMS Switch_Cutoff 19.03.26.xlsx")
OUTPUT_PATH = Path(r"D:\CIE\BMS\database\07_import_otb_switching_from_excel.sql")
SHEET_NAME = "Sheet1"
DEFAULT_USER = "Excel Import"


def normalize_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    return text or None


def sql_nvarchar(value) -> str:
    text = normalize_str(value)
    if text is None:
        return "NULL"
    return "N'" + text.replace("'", "''") + "'"


def sql_int(value) -> str:
    text = normalize_str(value)
    if text is None:
        return "NULL"
    return str(int(float(text)))


def sql_decimal(value) -> str:
    if value is None:
        return "0.00"
    if isinstance(value, (int, float)):
        return f"{float(value):.2f}"
    text = normalize_str(value)
    if text is None:
        return "0.00"
    return f"{float(text.replace(',', '')):.2f}"


def load_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [normalize_str(cell) or f"Column{idx+1}" for idx, cell in enumerate(rows[0])]
    data: list[dict[str, object]] = []
    for row in rows[1:]:
        if row is None:
            continue
        if all(cell is None or normalize_str(cell) is None for cell in row):
            continue
        data.append(dict(zip(headers, row)))
    return data


def get_switch_codes(from_year: object, from_month: object, to_year: object, to_month: object) -> tuple[str, str]:
    from_date = date(int(float(str(from_year))), int(float(str(from_month))), 1)
    to_date = date(int(float(str(to_year))), int(float(str(to_month))), 1)
    if from_date > to_date:
        return "G", "F"
    if from_date < to_date:
        return "I", "H"
    return "D", "C"


def chunked(items: list[str], size: int) -> Iterable[list[str]]:
    for idx in range(0, len(items), size):
        yield items[idx : idx + size]


def main() -> None:
    source_rows = load_rows(INPUT_PATH, SHEET_NAME)
    deduped_rows: OrderedDict[
        tuple[
            str | None,
            str | None,
            str | None,
            str | None,
            str | None,
            str | None,
            str | None,
            str,
            str | None,
            str | None,
            str | None,
            str | None,
            str | None,
            str,
            str | None,
            str | None,
            str,
            str | None,
        ],
        dict[str, object],
    ] = OrderedDict()
    duplicate_keys: list[tuple[str | None, ...]] = []

    for row in source_rows:
        from_code, to_code = get_switch_codes(row["Year"], row["Month"], row["To_Year"], row["To_Month"])
        amount_key = f"{float(str(row['Amount']).replace(',', '')):.2f}"
        key = (
            normalize_str(row.get("Year")),
            normalize_str(row.get("Month")),
            normalize_str(row.get("Company")),
            normalize_str(row.get("Category")),
            normalize_str(row.get("Segment")),
            normalize_str(row.get("Brand")),
            normalize_str(row.get("Vendor")),
            from_code,
            normalize_str(row.get("To_Year")),
            normalize_str(row.get("To_Month")),
            normalize_str(row.get("To_Company")),
            normalize_str(row.get("To_Category")),
            normalize_str(row.get("To_Segment")),
            to_code,
            normalize_str(row.get("To_Brand")),
            normalize_str(row.get("To_Vendor")),
            amount_key,
            normalize_str(row.get("Remark")),
        )
        if key in deduped_rows:
            duplicate_keys.append(key)
        deduped_rows[key] = row

    value_rows: list[str] = []
    for key, row in deduped_rows.items():
        from_code, to_code = key[7], key[13]
        value_rows.append(
            "("
            + ", ".join(
                [
                    sql_int(row.get("Year")),
                    sql_int(row.get("Month")),
                    sql_nvarchar(row.get("Company")),
                    sql_nvarchar(row.get("Category")),
                    sql_nvarchar(row.get("Segment")),
                    sql_nvarchar(row.get("Brand")),
                    sql_nvarchar(row.get("Vendor")),
                    sql_nvarchar(from_code),
                    sql_decimal(row.get("Amount")),
                    "0.00",
                    sql_int(row.get("To_Year")),
                    sql_int(row.get("To_Month")),
                    sql_nvarchar(row.get("To_Company")),
                    sql_nvarchar(row.get("To_Category")),
                    sql_nvarchar(row.get("To_Segment")),
                    sql_nvarchar(to_code),
                    sql_nvarchar(row.get("To_Brand")),
                    sql_nvarchar(row.get("To_Vendor")),
                    "N'Approved'",
                    sql_nvarchar(INPUT_PATH.stem),
                    sql_nvarchar(row.get("Remark")),
                    sql_nvarchar(DEFAULT_USER),
                    sql_nvarchar(DEFAULT_USER),
                ]
            )
            + ")"
        )

    lines = [
        "USE [BMS];",
        "GO",
        "",
        "SET NOCOUNT ON;",
        "SET XACT_ABORT ON;",
        "GO",
        "",
        f"/* Generated from: {INPUT_PATH} [{SHEET_NAME}] */",
        f"/* Source rows: {len(source_rows)} | Deduped rows used for merge: {len(value_rows)} */",
        "",
    ]

    if duplicate_keys:
        lines.append("/* Exact duplicate switching rows detected in source and collapsed to last occurrence:")
        for key in OrderedDict.fromkeys(duplicate_keys).keys():
            lines.append(
                "   "
                + f"Year={key[0]}, Month={key[1]}, Company={key[2]}, Category={key[3]}, Segment={key[4]}, Brand={key[5]}, Vendor={key[6]}, "
                + f"From={key[7]}, SwitchYear={key[8]}, SwitchMonth={key[9]}, SwitchCompany={key[10]}, SwitchCategory={key[11]}, SwitchSegment={key[12]}, "
                + f"To={key[13]}, SwitchBrand={key[14]}, SwitchVendor={key[15]}, Amount={key[16]}"
            )
        lines.append("*/")
        lines.append("")

    lines.extend(
        [
            "IF OBJECT_ID(N'tempdb..#OTB_Switching_Source', N'U') IS NOT NULL",
            "    DROP TABLE #OTB_Switching_Source;",
            "",
            "CREATE TABLE #OTB_Switching_Source",
            "(",
            "    [Year] int NOT NULL,",
            "    [Month] int NOT NULL,",
            "    Company nvarchar(20) NOT NULL,",
            "    Category nvarchar(20) NOT NULL,",
            "    Segment nvarchar(20) NOT NULL,",
            "    Brand nvarchar(30) NOT NULL,",
            "    Vendor nvarchar(30) NOT NULL,",
            "    [From] nvarchar(5) NOT NULL,",
            "    BudgetAmount decimal(18,2) NOT NULL,",
            "    [Release] decimal(18,2) NOT NULL,",
            "    SwitchYear int NULL,",
            "    SwitchMonth int NULL,",
            "    SwitchCompany nvarchar(20) NULL,",
            "    SwitchCategory nvarchar(20) NULL,",
            "    SwitchSegment nvarchar(20) NULL,",
            "    [To] nvarchar(5) NULL,",
            "    SwitchBrand nvarchar(30) NULL,",
            "    SwitchVendor nvarchar(30) NULL,",
            "    OTBStatus nvarchar(30) NOT NULL,",
            "    Batch nvarchar(50) NULL,",
            "    Remark nvarchar(500) NULL,",
            "    CreateBy nvarchar(100) NULL,",
            "    ActionBy nvarchar(100) NULL",
            ");",
            "",
        ]
    )

    for batch in chunked(value_rows, 500):
        lines.append("INSERT INTO #OTB_Switching_Source")
        lines.append("(")
        lines.append("    [Year], [Month], Company, Category, Segment, Brand, Vendor, [From], BudgetAmount, [Release],")
        lines.append("    SwitchYear, SwitchMonth, SwitchCompany, SwitchCategory, SwitchSegment, [To], SwitchBrand, SwitchVendor,")
        lines.append("    OTBStatus, Batch, Remark, CreateBy, ActionBy")
        lines.append(")")
        lines.append("VALUES")
        for idx, value_row in enumerate(batch):
            suffix = "," if idx < len(batch) - 1 else ";"
            lines.append(f"    {value_row}{suffix}")
        lines.append("")

    lines.extend(
        [
            "BEGIN TRANSACTION;",
            "",
            "MERGE dbo.OTB_Switching_Transaction AS T",
            "USING #OTB_Switching_Source AS S",
            "    ON T.[Year] = S.[Year]",
            "   AND T.[Month] = S.[Month]",
            "   AND T.Company = S.Company",
            "   AND T.Category = S.Category",
            "   AND T.Segment = S.Segment",
            "   AND T.Brand = S.Brand",
            "   AND T.Vendor = S.Vendor",
            "   AND T.[From] = S.[From]",
            "   AND T.BudgetAmount = S.BudgetAmount",
            "   AND ISNULL(T.SwitchYear, -1) = ISNULL(S.SwitchYear, -1)",
            "   AND ISNULL(T.SwitchMonth, -1) = ISNULL(S.SwitchMonth, -1)",
            "   AND ISNULL(T.SwitchCompany, N'') = ISNULL(S.SwitchCompany, N'')",
            "   AND ISNULL(T.SwitchCategory, N'') = ISNULL(S.SwitchCategory, N'')",
            "   AND ISNULL(T.SwitchSegment, N'') = ISNULL(S.SwitchSegment, N'')",
            "   AND ISNULL(T.[To], N'') = ISNULL(S.[To], N'')",
            "   AND ISNULL(T.SwitchBrand, N'') = ISNULL(S.SwitchBrand, N'')",
            "   AND ISNULL(T.SwitchVendor, N'') = ISNULL(S.SwitchVendor, N'')",
            "WHEN MATCHED THEN",
            "    UPDATE SET",
            "        T.[Release] = S.[Release],",
            "        T.OTBStatus = S.OTBStatus,",
            "        T.Batch = S.Batch,",
            "        T.Remark = S.Remark,",
            "        T.CreateBy = COALESCE(T.CreateBy, S.CreateBy),",
            "        T.ActionBy = S.ActionBy",
            "WHEN NOT MATCHED BY TARGET THEN",
            "    INSERT",
            "    (",
            "        [Year], [Month], Company, Category, Segment, Brand, Vendor, [From], BudgetAmount, [Release],",
            "        SwitchYear, SwitchMonth, SwitchCompany, SwitchCategory, SwitchSegment, [To], SwitchBrand, SwitchVendor,",
            "        OTBStatus, Batch, Remark, CreateBy, CreateDT, ActionBy",
            "    )",
            "    VALUES",
            "    (",
            "        S.[Year], S.[Month], S.Company, S.Category, S.Segment, S.Brand, S.Vendor, S.[From], S.BudgetAmount, S.[Release],",
            "        S.SwitchYear, S.SwitchMonth, S.SwitchCompany, S.SwitchCategory, S.SwitchSegment, S.[To], S.SwitchBrand, S.SwitchVendor,",
            "        S.OTBStatus, S.Batch, S.Remark, S.CreateBy, sysdatetime(), S.ActionBy",
            "    );",
            "",
            "COMMIT TRANSACTION;",
            "",
            "SELECT",
            "    COUNT(1) AS SourceRowCount",
            "FROM #OTB_Switching_Source;",
            "",
            "PRINT N'OTB_Switching_Transaction import merge completed.';",
            "GO",
            "",
        ]
    )

    OUTPUT_PATH.write_text("\n".join(lines), encoding="utf-8")
    print(f"Generated: {OUTPUT_PATH}")
    print(f"Source rows: {len(source_rows)}")
    print(f"Deduped rows used for merge: {len(value_rows)}")


if __name__ == "__main__":
    main()
