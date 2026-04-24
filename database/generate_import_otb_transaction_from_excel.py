from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


INPUT_PATH = Path(r"C:\Users\60893\Downloads\BMS_Approved_OTB 2026.xlsx")
OUTPUT_PATH = Path(r"D:\CIE\BMS\database\06_import_otb_transaction_from_excel.sql")
SHEET_NAME = "Data"

MONTH_MAP = {
    "1": 1,
    "01": 1,
    "jan": 1,
    "january": 1,
    "2": 2,
    "02": 2,
    "feb": 2,
    "february": 2,
    "3": 3,
    "03": 3,
    "mar": 3,
    "march": 3,
    "4": 4,
    "04": 4,
    "apr": 4,
    "april": 4,
    "5": 5,
    "05": 5,
    "may": 5,
    "6": 6,
    "06": 6,
    "jun": 6,
    "june": 6,
    "7": 7,
    "07": 7,
    "jul": 7,
    "july": 7,
    "8": 8,
    "08": 8,
    "aug": 8,
    "august": 8,
    "9": 9,
    "09": 9,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "10": 10,
    "oct": 10,
    "october": 10,
    "11": 11,
    "nov": 11,
    "november": 11,
    "12": 12,
    "dec": 12,
    "december": 12,
}


def normalize_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    return text or None


def sql_nvarchar(value) -> str:
    text = normalize_str(value)
    if text is None:
        return "NULL"
    return "N'" + text.replace("'", "''") + "'"


def sql_int(value) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value))
    text = normalize_str(value)
    if text is None:
        return "NULL"
    return str(int(float(text)))


def sql_decimal(value) -> str:
    if value is None:
        return "0"
    if isinstance(value, (int, float)):
        return f"{float(value):.2f}"
    text = normalize_str(value)
    if text is None:
        return "0"
    text = text.replace(",", "")
    return f"{float(text):.2f}"


def parse_datetime(value) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = normalize_str(value)
    if text is None:
        return None
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def sql_datetime(value) -> str:
    dt = parse_datetime(value)
    if dt is None:
        return "NULL"
    return f"CAST('{dt:%Y-%m-%d %H:%M:%S}' AS datetime2(0))"


def month_to_int(value) -> int:
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = normalize_str(value)
    if text is None:
        raise ValueError("Month is required.")
    month = MONTH_MAP.get(text.lower())
    if month is None:
        raise ValueError(f"Unsupported month value: {text}")
    return month


def load_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [normalize_str(cell) or f"Column{idx+1}" for idx, cell in enumerate(rows[0])]
    data: list[dict[str, object]] = []
    for row in rows[1:]:
        if row is None:
            continue
        if all(cell is None or normalize_str(cell) is None for cell in row):
            continue
        data.append(dict(zip(headers, row)))
    return data


def chunked(items: list[str], size: int) -> Iterable[list[str]]:
    for idx in range(0, len(items), size):
        yield items[idx : idx + size]


def main() -> None:
    rows = load_rows(INPUT_PATH, SHEET_NAME)
    value_rows: list[str] = []

    for row in rows:
        month_no = month_to_int(row.get("Month"))
        action_by = normalize_str(row.get("Approved By")) or normalize_str(row.get("Create By"))
        value_rows.append(
            "("
            + ", ".join(
                [
                    sql_datetime(row.get("Create date")),
                    sql_nvarchar(row.get("Type")),
                    sql_int(row.get("Year")),
                    sql_int(month_no),
                    sql_nvarchar(row.get("Category")),
                    sql_nvarchar(row.get("Category name")),
                    sql_nvarchar(row.get("Company")),
                    sql_nvarchar(row.get("Segment")),
                    sql_nvarchar(row.get("Segment name")),
                    sql_nvarchar(row.get("Brand")),
                    sql_nvarchar(row.get("Brand name")),
                    sql_nvarchar(row.get("Vendor")),
                    sql_nvarchar(row.get("Vendor name")),
                    sql_decimal(row.get("Amount (THB)")),
                    sql_decimal(row.get("Revised Diff")),
                    sql_nvarchar(row.get("Remark")),
                    sql_nvarchar(row.get("Status")),
                    sql_datetime(row.get("Approved Date")),
                    sql_nvarchar(action_by),
                    sql_nvarchar(row.get("SAP Status")),
                    sql_nvarchar(row.get("Version")),
                ]
            )
            + ")"
        )

    lines = [
        "USE [BMS];",
        "GO",
        "",
        "SET NOCOUNT ON;",
        "SET XACT_ABORT ON;",
        "GO",
        "",
        f"/* Generated from: {INPUT_PATH} [{SHEET_NAME}] */",
        f"/* Source rows: {len(value_rows)} */",
        "",
        "IF OBJECT_ID(N'tempdb..#OTB_Transaction_Source', N'U') IS NOT NULL",
        "    DROP TABLE #OTB_Transaction_Source;",
        "",
        "CREATE TABLE #OTB_Transaction_Source",
        "(",
        "    CreateDate datetime2(0) NULL,",
        "    [Type] nvarchar(20) NOT NULL,",
        "    [Year] int NOT NULL,",
        "    [Month] int NOT NULL,",
        "    Category nvarchar(20) NOT NULL,",
        "    CategoryName nvarchar(200) NULL,",
        "    Company nvarchar(20) NOT NULL,",
        "    Segment nvarchar(20) NOT NULL,",
        "    SegmentName nvarchar(200) NULL,",
        "    Brand nvarchar(30) NOT NULL,",
        "    BrandName nvarchar(200) NULL,",
        "    Vendor nvarchar(30) NOT NULL,",
        "    VendorName nvarchar(200) NULL,",
        "    Amount decimal(18,2) NOT NULL,",
        "    RevisedDiff decimal(18,2) NULL,",
        "    Remark nvarchar(500) NULL,",
        "    OTBStatus nvarchar(30) NOT NULL,",
        "    ApprovedDate datetime2(0) NULL,",
        "    ActionBy nvarchar(100) NULL,",
        "    SAPStatus nvarchar(10) NULL,",
        "    [Version] nvarchar(20) NOT NULL",
        ");",
        "",
    ]

    for batch in chunked(value_rows, 500):
        lines.append("INSERT INTO #OTB_Transaction_Source")
        lines.append("(")
        lines.append("    CreateDate, [Type], [Year], [Month], Category, CategoryName, Company, Segment, SegmentName,")
        lines.append("    Brand, BrandName, Vendor, VendorName, Amount, RevisedDiff, Remark, OTBStatus, ApprovedDate,")
        lines.append("    ActionBy, SAPStatus, [Version]")
        lines.append(")")
        lines.append("VALUES")
        for idx, value_row in enumerate(batch):
            suffix = "," if idx < len(batch) - 1 else ";"
            lines.append(f"    {value_row}{suffix}")
        lines.append("")

    lines.extend(
        [
            "BEGIN TRANSACTION;",
            "",
            "MERGE dbo.OTB_Transaction AS T",
            "USING #OTB_Transaction_Source AS S",
            "    ON T.[Type] = S.[Type]",
            "   AND T.[Year] = S.[Year]",
            "   AND T.[Month] = S.[Month]",
            "   AND T.Category = S.Category",
            "   AND T.Company = S.Company",
            "   AND T.Segment = S.Segment",
            "   AND T.Brand = S.Brand",
            "   AND T.Vendor = S.Vendor",
            "   AND T.[Version] = S.[Version]",
            "WHEN MATCHED THEN",
            "    UPDATE SET",
            "        T.CreateDate = ISNULL(S.CreateDate, T.CreateDate),",
            "        T.CategoryName = S.CategoryName,",
            "        T.SegmentName = S.SegmentName,",
            "        T.BrandName = S.BrandName,",
            "        T.VendorName = S.VendorName,",
            "        T.Amount = S.Amount,",
            "        T.RevisedDiff = S.RevisedDiff,",
            "        T.Remark = S.Remark,",
            "        T.OTBStatus = S.OTBStatus,",
            "        T.ApprovedDate = S.ApprovedDate,",
            "        T.ActionBy = S.ActionBy,",
            "        T.SAPStatus = S.SAPStatus",
            "WHEN NOT MATCHED BY TARGET THEN",
            "    INSERT",
            "    (",
            "        CreateDate, [Type], [Year], [Month], Category, CategoryName, Company, Segment, SegmentName,",
            "        Brand, BrandName, Vendor, VendorName, Amount, RevisedDiff, Remark, OTBStatus, ApprovedDate,",
            "        SAPDate, ActionBy, DraftID, SAPStatus, SAPErrorMessage, [Version]",
            "    )",
            "    VALUES",
            "    (",
            "        ISNULL(S.CreateDate, sysdatetime()), S.[Type], S.[Year], S.[Month], S.Category, S.CategoryName, S.Company, S.Segment, S.SegmentName,",
            "        S.Brand, S.BrandName, S.Vendor, S.VendorName, S.Amount, S.RevisedDiff, S.Remark, S.OTBStatus, S.ApprovedDate,",
            "        NULL, S.ActionBy, NULL, S.SAPStatus, NULL, S.[Version]",
            "    );",
            "",
            "COMMIT TRANSACTION;",
            "",
            "SELECT",
            "    COUNT(1) AS SourceRowCount,",
            "    COUNT(DISTINCT CONCAT([Type], N'|', [Year], N'|', [Month], N'|', Category, N'|', Company, N'|', Segment, N'|', Brand, N'|', Vendor, N'|', [Version])) AS DistinctBusinessKeyCount",
            "FROM #OTB_Transaction_Source;",
            "",
            "PRINT N'OTB_Transaction import merge completed.';",
            "GO",
            "",
        ]
    )

    OUTPUT_PATH.write_text("\n".join(lines), encoding="utf-8")
    print(f"Generated: {OUTPUT_PATH}")
    print(f"Source rows: {len(value_rows)}")


if __name__ == "__main__":
    main()
