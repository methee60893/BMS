from __future__ import annotations

from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


BASE_DIR = Path(r"D:\202603_Deaktop\BMS")
OUTPUT_PATH = Path(r"D:\CIE\BMS\database\04_seed_master_data_template.sql")
VERIFY_SQL_PATH = Path(r"D:\CIE\BMS\database\05_post_deploy_verify.sql")
SUPPLEMENTAL_MASTER_PATH = Path(r"C:\Users\60893\Downloads\MS_BMS_OLD_Update 07.04.26.xlsx")
SUPPLEMENTAL_VENDOR_PATH = Path(r"C:\Users\60893\Downloads\MasterVendor.xlsx")

MENU_ROWS = [
    (1, "Dashboard", "dashboard.aspx", "Main", 1, 1),
    (2, "Draft OTB", "draftOTB.aspx", "OTB", 2, 1),
    (3, "Approved OTB", "approvedOTB.aspx", "OTB", 3, 1),
    (4, "Create Draft PO", "createDraftPO.aspx", "PO", 4, 1),
    (5, "Create OTB Switching", "createOTBswitching.aspx", "OTB", 5, 1),
    (6, "Actual PO", "actualPO.aspx", "PO", 6, 1),
    (7, "Draft PO", "draftPO.aspx", "PO", 7, 1),
    (8, "Master Brand", "master_brand.aspx", "Master", 8, 1),
    (9, "Master Category", "master_category.aspx", "Master", 9, 1),
    (10, "Master Vendor", "master_vendor.aspx", "Master", 10, 1),
    (11, "Match Actual PO", "matchActualPO.aspx", "PO", 11, 1),
    (12, "OTB Remaining", "otbRemaining.aspx", "OTB", 12, 1),
    (13, "Admin Match PO", "admin_matchPO.aspx", "PO", 13, 1),
    (14, "Switching Transaction", "transactionOTBSwitching.aspx", "OTB", 14, 1),
    (15, "Manage Users", "manage_users.aspx", "Admin", 15, 1),
    (16, "Login", "default.aspx", "System", 16, 1),
]

ADDITIONAL_PERMISSION_ROWS = [
    (31, 1, 14, 1, 1, 1, 1),
    (32, 3, 14, 1, 1, 1, 1),
    (33, 1, 15, 1, 1, 1, 1),
    (34, 1, 16, 1, 1, 1, 1),
    (35, 2, 16, 1, 0, 0, 0),
    (36, 3, 16, 1, 0, 0, 0),
]


def normalize_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def sql_nvarchar(value) -> str:
    text = normalize_str(value)
    if text is None:
        return "NULL"
    return "N'" + text.replace("'", "''") + "'"


def sql_number(value) -> str:
    text = normalize_str(value)
    return "NULL" if text is None else text


def sql_bit(value) -> str:
    text = normalize_str(value)
    if text is None:
        return "0"
    return "1" if text.lower() in {"1", "true", "yes", "active"} else "0"


def sql_datetime(value) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, datetime):
        return f"CAST('{value:%Y-%m-%d %H:%M:%S}' AS datetime2(0))"
    text = normalize_str(value)
    if text is None:
        return "NULL"
    try:
        dt = datetime.fromisoformat(text)
    except ValueError:
        return "NULL"
    return f"CAST('{dt:%Y-%m-%d %H:%M:%S}' AS datetime2(0))"


def load_sheet(path: Path, sheet_name: str) -> list[OrderedDict[str, object]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [normalize_str(cell) or f"Column{idx+1}" for idx, cell in enumerate(rows[0])]
    data: list[OrderedDict[str, object]] = []
    for row in rows[1:]:
        if row is None:
            continue
        if all(cell is None or normalize_str(cell) is None for cell in row):
            continue
        data.append(OrderedDict(zip(headers, row)))
    return data


def load_optional_sheet(path: Path, sheet_name: str) -> list[OrderedDict[str, object]]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_str(cell) or f"Column{idx+1}" for idx, cell in enumerate(rows[0])]
    data: list[OrderedDict[str, object]] = []
    for row in rows[1:]:
        if row is None:
            continue
        if all(cell is None or normalize_str(cell) is None for cell in row):
            continue
        data.append(OrderedDict(zip(headers, row)))
    return data


def append_merge(
    lines: list[str],
    target: str,
    source_columns: list[str],
    rows: Iterable[str],
    on_clause: str,
    update_clause: str | None,
    insert_columns: list[str],
    insert_values: list[str],
) -> None:
    row_list = list(rows)
    if not row_list:
        return
    lines.append(f"MERGE {target} AS T")
    lines.append("USING")
    lines.append("(")
    lines.append("    VALUES")
    for idx, row in enumerate(row_list):
        suffix = "," if idx < len(row_list) - 1 else ""
        lines.append(f"        {row}{suffix}")
    lines.append(") AS S (" + ", ".join(source_columns) + ")")
    lines.append(f"ON {on_clause}")
    if update_clause:
        lines.append("WHEN MATCHED THEN")
        lines.append("    UPDATE SET")
        lines.append(f"        {update_clause}")
    lines.append("WHEN NOT MATCHED BY TARGET THEN")
    lines.append("    INSERT (" + ", ".join(insert_columns) + ")")
    lines.append("    VALUES (" + ", ".join(insert_values) + ");")
    lines.append("GO")
    lines.append("")


def build_verify_sql(expected_counts: list[tuple[str, int]]) -> str:
    lines = [
        "USE [BMS];",
        "GO",
        "",
        "SET NOCOUNT ON;",
        "GO",
        "",
        "DECLARE @Expected TABLE",
        "(",
        "    TableName sysname NOT NULL,",
        "    ExpectedCount int NOT NULL",
        ");",
        "",
        "DECLARE @ActualCounts TABLE",
        "(",
        "    TableName sysname NOT NULL,",
        "    ActualCount int NOT NULL",
        ");",
        "",
        "INSERT INTO @Expected (TableName, ExpectedCount)",
        "VALUES",
    ]

    for idx, (table_name, count) in enumerate(expected_counts):
        suffix = "," if idx < len(expected_counts) - 1 else ";"
        lines.append(f"    (N'{table_name}', {count}){suffix}")

    lines.extend(
        [
            "",
            "INSERT INTO @ActualCounts (TableName, ActualCount)",
            "SELECT N'MS_Vendor' AS TableName, COUNT(1) AS ActualCount FROM dbo.MS_Vendor",
            "UNION ALL SELECT N'MS_Brand', COUNT(1) FROM dbo.MS_Brand",
            "UNION ALL SELECT N'MS_Category', COUNT(1) FROM dbo.MS_Category",
            "UNION ALL SELECT N'MS_User', COUNT(1) FROM dbo.MS_User",
            "UNION ALL SELECT N'MS_Role', COUNT(1) FROM dbo.MS_Role",
            "UNION ALL SELECT N'MS_Menu', COUNT(1) FROM dbo.MS_Menu",
            "UNION ALL SELECT N'Map_User_Role', COUNT(1) FROM dbo.Map_User_Role",
            "UNION ALL SELECT N'Map_Role_Permission', COUNT(1) FROM dbo.Map_Role_Permission;",
            "",
            "SELECT",
            "    e.TableName,",
            "    e.ExpectedCount,",
            "    a.ActualCount,",
            "    CASE WHEN e.ExpectedCount = a.ActualCount THEN N'PASS' ELSE N'FAIL' END AS VerifyStatus",
            "FROM @Expected e",
            "LEFT JOIN @ActualCounts a",
            "    ON e.TableName = a.TableName",
            "ORDER BY e.TableName;",
            "",
            "SELECT",
            "    CASE",
            "        WHEN EXISTS",
            "        (",
            "            SELECT 1",
            "            FROM @Expected e",
            "            LEFT JOIN @ActualCounts a",
            "                ON e.TableName = a.TableName",
            "            WHERE ISNULL(a.ActualCount, -1) <> e.ExpectedCount",
            "        )",
            "        THEN N'FAIL'",
            "        ELSE N'PASS'",
            "    END AS OverallStatus;",
            "GO",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    vendor_rows_raw = load_sheet(BASE_DIR / "1.Master Vendor_BMS_v2.xlsx", "vendor")
    brand_rows = load_sheet(BASE_DIR / "2.Master Brand_BMS_V1.xlsx", "brand")
    category_rows = load_sheet(BASE_DIR / "3.Master category_BMS_24.11.25.xlsx", "Sheet1")
    user_rows = load_sheet(BASE_DIR / "Data_KBMS_User_Role.xlsx", "MS_User")
    role_rows = load_sheet(BASE_DIR / "Data_KBMS_User_Role.xlsx", "MS_Role")
    user_role_rows = load_sheet(BASE_DIR / "Data_KBMS_User_Role.xlsx", "Map_User_Role")
    role_perm_rows = load_sheet(BASE_DIR / "Data_KBMS_User_Role.xlsx", " Map_Role_Permission")
    supplemental_segment_rows = load_optional_sheet(SUPPLEMENTAL_MASTER_PATH, "Segment")
    supplemental_brand_rows = load_optional_sheet(SUPPLEMENTAL_MASTER_PATH, "Brand_update07.4.26")
    supplemental_vendor_rows = load_optional_sheet(SUPPLEMENTAL_VENDOR_PATH, "MasterVendor_20260319151554")
    existing_permission_ids = {int(row["PermissionID"]) for row in role_perm_rows}
    for permission_id, role_id, menu_id, can_view, can_edit, can_delete, can_approve in ADDITIONAL_PERMISSION_ROWS:
        if permission_id not in existing_permission_ids:
            role_perm_rows.append(
                OrderedDict(
                    [
                        ("PermissionID", permission_id),
                        ("RoleID", role_id),
                        ("MenuID", menu_id),
                        ("CanView", can_view),
                        ("CanEdit", can_edit),
                        ("CanDelete", can_delete),
                        ("CanApprove", can_approve),
                    ]
                )
            )

    brand_map: OrderedDict[str, OrderedDict[str, object]] = OrderedDict()
    for row in brand_rows:
        brand_code = normalize_str(row.get("Brand Code"))
        if brand_code:
            brand_map[brand_code] = row
    added_brand_codes: list[str] = []
    for row in supplemental_brand_rows:
        brand_code = normalize_str(row.get("Brand ID"))
        if brand_code and brand_code not in brand_map:
            brand_map[brand_code] = OrderedDict(
                [
                    ("Brand Code", brand_code),
                    ("Brand Name", normalize_str(row.get("Brand Description")) or brand_code),
                    ("Status", "Active"),
                ]
            )
            added_brand_codes.append(brand_code)
    brand_rows = list(brand_map.values())

    vendor_map: OrderedDict[tuple[str | None, str | None, str | None], OrderedDict[str, object]] = OrderedDict()
    duplicate_vendor_keys: list[tuple[str | None, str | None, str | None]] = []
    for row in vendor_rows_raw:
        key = (
            normalize_str(row.get("Vendor Code")),
            normalize_str(row.get("Segment Code")),
            normalize_str(row.get("Payment Term Code")),
        )
        if key in vendor_map:
            duplicate_vendor_keys.append(key)
        vendor_map[key] = row

    segment_map: OrderedDict[str, str] = OrderedDict()
    segment_active_map: OrderedDict[str, object] = OrderedDict()
    ccy_set: OrderedDict[str, None] = OrderedDict()
    for row in vendor_map.values():
        segment_code = normalize_str(row.get("Segment Code"))
        segment_name = normalize_str(row.get("Segment"))
        ccy = normalize_str(row.get("CCY"))
        if segment_code and segment_code not in segment_map:
            segment_map[segment_code] = segment_name or segment_code
            segment_active_map[segment_code] = row.get("isActive", 1)
        if ccy and ccy not in ccy_set:
            ccy_set[ccy] = None

    added_segment_codes: list[str] = []
    for row in supplemental_segment_rows:
        segment_code = normalize_str(row.get("SegmentCode"))
        segment_name = normalize_str(row.get("SegmentName"))
        if segment_code and segment_code not in segment_map:
            segment_map[segment_code] = segment_name or segment_code
            segment_active_map[segment_code] = row.get("isActive") if row.get("isActive") is not None else 1
            added_segment_codes.append(segment_code)

    base_vendor_keys = set(vendor_map.keys())
    added_vendor_keys: list[tuple[str | None, str | None, str | None]] = []
    duplicate_supplemental_vendor_keys: list[tuple[str | None, str | None, str | None]] = []
    supplemental_vendor_seen: set[tuple[str | None, str | None, str | None]] = set()
    for row in supplemental_vendor_rows:
        normalized_row = OrderedDict(
            [
                ("Vendor Code", normalize_str(row.get("Vendor Code"))),
                ("Vendor Name", normalize_str(row.get("Vendor Name"))),
                ("CCY", normalize_str(row.get("CCY"))),
                ("Payment Term Code", normalize_str(row.get("Payment Term Code"))),
                ("Payment Term", normalize_str(row.get("Payment Term"))),
                ("Segment Code", normalize_str(row.get("Segment Code"))),
                ("Segment", normalize_str(row.get("Segment"))),
                ("Incoterm", normalize_str(row.get("Incoterm"))),
                ("Status", normalize_str(row.get("Status"))),
            ]
        )
        key = (
            normalized_row["Vendor Code"],
            normalized_row["Segment Code"],
            normalized_row["Payment Term Code"],
        )
        if not key[0]:
            continue
        if key in base_vendor_keys:
            continue
        if key in supplemental_vendor_seen:
            duplicate_supplemental_vendor_keys.append(key)
            vendor_map[key] = normalized_row
            continue
        supplemental_vendor_seen.add(key)
        vendor_map[key] = normalized_row
        added_vendor_keys.append(key)
        segment_code = normalized_row["Segment Code"]
        segment_name = normalized_row["Segment"]
        ccy = normalized_row["CCY"]
        if segment_code and segment_code not in segment_map:
            segment_map[segment_code] = segment_name or segment_code
            segment_active_map[segment_code] = 1
            added_segment_codes.append(segment_code)
        if ccy and ccy not in ccy_set:
            ccy_set[ccy] = None

    vendor_rows = list(vendor_map.values())

    lines: list[str] = [
        "USE [BMS];",
        "GO",
        "",
        "SET NOCOUNT ON;",
        "GO",
        "",
        "/* Generated from provided Excel files using openpyxl */",
        f"/* Vendor source rows: {len(vendor_rows_raw)} | Vendor unique rows used for seed: {len(vendor_rows)} */",
        f"/* Supplemental additions: Segment +{len(added_segment_codes)} | Brand +{len(added_brand_codes)} | Vendor +{len(OrderedDict.fromkeys(added_vendor_keys))} */",
        "",
    ]
    if duplicate_vendor_keys:
        unique_dup_keys = list(OrderedDict.fromkeys(duplicate_vendor_keys).keys())
        lines.append("/* Duplicate VendorCode+SegmentCode+PaymentTermCode detected in source and collapsed to last occurrence:")
        for vendor_code, segment_code, payment_term_code in unique_dup_keys:
            lines.append(
                f"   VendorCode={vendor_code or 'NULL'}, SegmentCode={segment_code or 'NULL'}, PaymentTermCode={payment_term_code or 'NULL'}"
            )
        lines.append("*/")
        lines.append("")
    if added_segment_codes:
        lines.append("/* Added SegmentCode from supplemental sheet:")
        for segment_code in OrderedDict.fromkeys(added_segment_codes).keys():
            lines.append(f"   SegmentCode={segment_code}")
        lines.append("*/")
        lines.append("")
    if duplicate_supplemental_vendor_keys:
        lines.append("/* Duplicate VendorCode+SegmentCode+PaymentTermCode detected in supplemental vendor file and last occurrence was kept:")
        for vendor_code, segment_code, payment_term_code in OrderedDict.fromkeys(duplicate_supplemental_vendor_keys).keys():
            lines.append(
                f"   VendorCode={vendor_code or 'NULL'}, SegmentCode={segment_code or 'NULL'}, PaymentTermCode={payment_term_code or 'NULL'}"
            )
        lines.append("*/")
        lines.append("")

    append_merge(
        lines,
        "dbo.MS_Month",
        ["month_code", "month_name_sh"],
        [
            "(1, N'Jan')",
            "(2, N'Feb')",
            "(3, N'Mar')",
            "(4, N'Apr')",
            "(5, N'May')",
            "(6, N'Jun')",
            "(7, N'Jul')",
            "(8, N'Aug')",
            "(9, N'Sep')",
            "(10, N'Oct')",
            "(11, N'Nov')",
            "(12, N'Dec')",
        ],
        "T.month_code = S.month_code",
        "T.month_name_sh = S.month_name_sh",
        ["month_code", "month_name_sh"],
        ["S.month_code", "S.month_name_sh"],
    )

    append_merge(
        lines,
        "dbo.MS_Year",
        ["Year_Kept"],
        ["(2026)", "(2027)", "(2028)"],
        "T.Year_Kept = S.Year_Kept",
        None,
        ["Year_Kept"],
        ["S.Year_Kept"],
    )

    append_merge(
        lines,
        "dbo.MS_CCY",
        ["CCY_Code", "CCY_Name", "isActive"],
        [f"({sql_nvarchar(ccy)}, {sql_nvarchar(ccy)}, 1)" for ccy in ccy_set.keys()],
        "T.CCY_Code = S.CCY_Code",
        "T.CCY_Name = S.CCY_Name, T.isActive = S.isActive",
        ["CCY_Code", "CCY_Name", "isActive"],
        ["S.CCY_Code", "S.CCY_Name", "S.isActive"],
    )

    append_merge(
        lines,
        "dbo.MS_Company",
        ["CompanyCode", "CompanyNameShort"],
        [
            "(N'1000', N'KPC')",
            "(N'2000', N'KPD')",
            "(N'3000', N'KPT')",
        ],
        "T.CompanyCode = S.CompanyCode",
        "T.CompanyNameShort = S.CompanyNameShort",
        ["CompanyCode", "CompanyNameShort"],
        ["S.CompanyCode", "S.CompanyNameShort"],
    )

    append_merge(
        lines,
        "dbo.MS_Segment",
        ["SegmentCode", "SegmentName", "isActive"],
        [
            f"({sql_nvarchar(code)}, {sql_nvarchar(name)}, {sql_bit(segment_active_map.get(code, 1))})"
            for code, name in segment_map.items()
        ],
        "T.SegmentCode = S.SegmentCode",
        "T.SegmentName = S.SegmentName, T.isActive = S.isActive",
        ["SegmentCode", "SegmentName", "isActive"],
        ["S.SegmentCode", "S.SegmentName", "S.isActive"],
    )

    append_merge(
        lines,
        "dbo.MS_Category",
        ["Cate", "Category", "isActive"],
        [
            f"({sql_nvarchar(row['Category Code'])}, {sql_nvarchar(row['Category Name'])}, {sql_bit(row['isActive'])})"
            for row in category_rows
        ],
        "T.Cate = S.Cate",
        "T.Category = S.Category, T.isActive = S.isActive",
        ["Cate", "Category", "isActive"],
        ["S.Cate", "S.Category", "S.isActive"],
    )

    append_merge(
        lines,
        "dbo.MS_Brand",
        ["[Brand Code]", "[Brand Name]", "isActive"],
        [
            f"({sql_nvarchar(row['Brand Code'])}, {sql_nvarchar(row['Brand Name'])}, {sql_bit(row['Status'])})"
            for row in brand_rows
        ],
        "T.[Brand Code] = S.[Brand Code]",
        "T.[Brand Name] = S.[Brand Name], T.isActive = S.isActive",
        ["[Brand Code]", "[Brand Name]", "isActive"],
        ["S.[Brand Code]", "S.[Brand Name]", "S.isActive"],
    )

    append_merge(
        lines,
        "dbo.MS_Version",
        ["VersionCode", "OTBTypeCode", "Seq", "isActive"],
        [
            "(N'A1', N'Original', 1, 1)",
            "(N'R1', N'Revise', 1, 1)",
            "(N'R2', N'Revise', 2, 1)",
            "(N'R3', N'Revise', 3, 1)",
        ],
        "T.VersionCode = S.VersionCode AND T.OTBTypeCode = S.OTBTypeCode",
        "T.Seq = S.Seq, T.isActive = S.isActive",
        ["VersionCode", "OTBTypeCode", "Seq", "isActive"],
        ["S.VersionCode", "S.OTBTypeCode", "S.Seq", "S.isActive"],
    )

    append_merge(
        lines,
        "dbo.MS_Vendor",
        ["VendorCode", "Vendor", "CCY", "PaymentTermCode", "PaymentTerm", "SegmentCode", "Segment", "Incoterm", "isActive"],
        [
            "("
            + ", ".join(
                [
                    sql_nvarchar(row["Vendor Code"]),
                    sql_nvarchar(row["Vendor Name"]),
                    sql_nvarchar(row["CCY"]),
                    sql_nvarchar(row["Payment Term Code"]),
                    sql_nvarchar(row["Payment Term"]),
                    sql_nvarchar(row["Segment Code"]),
                    sql_nvarchar(row["Segment"]),
                    sql_nvarchar(row["Incoterm"]),
                    sql_bit(row["Status"]),
                ]
            )
            + ")"
            for row in vendor_rows
        ],
        "T.VendorCode = S.VendorCode AND ISNULL(T.SegmentCode, N'') = ISNULL(S.SegmentCode, N'') AND ISNULL(T.PaymentTermCode, N'') = ISNULL(S.PaymentTermCode, N'')",
        "T.Vendor = S.Vendor, T.CCY = S.CCY, T.PaymentTermCode = S.PaymentTermCode, T.PaymentTerm = S.PaymentTerm, T.Segment = S.Segment, T.Incoterm = S.Incoterm, T.isActive = S.isActive",
        ["VendorCode", "Vendor", "CCY", "PaymentTermCode", "PaymentTerm", "SegmentCode", "Segment", "Incoterm", "isActive"],
        ["S.VendorCode", "S.Vendor", "S.CCY", "S.PaymentTermCode", "S.PaymentTerm", "S.SegmentCode", "S.Segment", "S.Incoterm", "S.isActive"],
    )

    append_merge(
        lines,
        "dbo.MS_User",
        ["UserID", "Username", "FullName", "Email", "IsActive", "LastLoginDate", "CreateDate"],
        [
            "("
            + ", ".join(
                [
                    sql_number(row["UserID"]),
                    sql_nvarchar(row["Username"]),
                    sql_nvarchar(row["FullName"]),
                    sql_nvarchar(row["Email"]),
                    sql_bit(row["IsActive"]),
                    sql_datetime(row["LastLoginDate"]),
                    sql_datetime(row["CreateDate"]),
                ]
            )
            + ")"
            for row in user_rows
        ],
        "T.UserID = S.UserID",
        "T.Username = S.Username, T.FullName = S.FullName, T.Email = S.Email, T.IsActive = S.IsActive, T.LastLoginDate = S.LastLoginDate, T.CreateDate = S.CreateDate",
        ["UserID", "Username", "FullName", "Email", "IsActive", "LastLoginDate", "CreateDate"],
        ["S.UserID", "S.Username", "S.FullName", "S.Email", "S.IsActive", "S.LastLoginDate", "S.CreateDate"],
    )

    append_merge(
        lines,
        "dbo.MS_Role",
        ["RoleID", "RoleName", "Description"],
        [
            f"({sql_number(row['RoleID'])}, {sql_nvarchar(row['RoleName'])}, {sql_nvarchar(row['Description'])})"
            for row in role_rows
        ],
        "T.RoleID = S.RoleID",
        "T.RoleName = S.RoleName, T.Description = S.Description",
        ["RoleID", "RoleName", "Description"],
        ["S.RoleID", "S.RoleName", "S.Description"],
    )

    append_merge(
        lines,
        "dbo.MS_Menu",
        ["MenuID", "MenuName", "PageUrl", "MenuGroup", "SortOrder", "IsActive"],
        [
            "("
            + ", ".join(
                [
                    sql_number(menu_id),
                    sql_nvarchar(menu_name),
                    sql_nvarchar(page_url),
                    sql_nvarchar(menu_group),
                    sql_number(sort_order),
                    sql_bit(is_active),
                ]
            )
            + ")"
            for menu_id, menu_name, page_url, menu_group, sort_order, is_active in MENU_ROWS
        ],
        "T.MenuID = S.MenuID",
        "T.MenuName = S.MenuName, T.PageUrl = S.PageUrl, T.MenuGroup = S.MenuGroup, T.SortOrder = S.SortOrder, T.IsActive = S.IsActive",
        ["MenuID", "MenuName", "PageUrl", "MenuGroup", "SortOrder", "IsActive"],
        ["S.MenuID", "S.MenuName", "S.PageUrl", "S.MenuGroup", "S.SortOrder", "S.IsActive"],
    )

    append_merge(
        lines,
        "dbo.Map_User_Role",
        ["MapID", "UserID", "RoleID"],
        [
            f"({sql_number(row['MapID'])}, {sql_number(row['UserID'])}, {sql_number(row['RoleID'])})"
            for row in user_role_rows
        ],
        "T.MapID = S.MapID",
        "T.UserID = S.UserID, T.RoleID = S.RoleID",
        ["MapID", "UserID", "RoleID"],
        ["S.MapID", "S.UserID", "S.RoleID"],
    )

    append_merge(
        lines,
        "dbo.Map_Role_Permission",
        ["PermissionID", "RoleID", "MenuID", "CanView", "CanEdit", "CanDelete", "CanApprove"],
        [
            "("
            + ", ".join(
                [
                    sql_number(row["PermissionID"]),
                    sql_number(row["RoleID"]),
                    sql_number(row["MenuID"]),
                    sql_bit(row["CanView"]),
                    sql_bit(row["CanEdit"]),
                    sql_bit(row["CanDelete"]),
                    sql_bit(row["CanApprove"]),
                ]
            )
            + ")"
            for row in role_perm_rows
        ],
        "T.PermissionID = S.PermissionID",
        "T.RoleID = S.RoleID, T.MenuID = S.MenuID, T.CanView = S.CanView, T.CanEdit = S.CanEdit, T.CanDelete = S.CanDelete, T.CanApprove = S.CanApprove",
        ["PermissionID", "RoleID", "MenuID", "CanView", "CanEdit", "CanDelete", "CanApprove"],
        ["S.PermissionID", "S.RoleID", "S.MenuID", "S.CanView", "S.CanEdit", "S.CanDelete", "S.CanApprove"],
    )

    lines.append("PRINT N'Master data seed generated from Excel completed.';")
    lines.append("GO")
    lines.append("")

    OUTPUT_PATH.write_text("\n".join(lines), encoding="utf-8")
    VERIFY_SQL_PATH.write_text(
        build_verify_sql(
            [
                ("MS_Vendor", len(vendor_rows)),
                ("MS_Brand", len(brand_rows)),
                ("MS_Category", len(category_rows)),
                ("MS_User", len(user_rows)),
                ("MS_Role", len(role_rows)),
                ("MS_Menu", len(MENU_ROWS)),
                ("Map_User_Role", len(user_role_rows)),
                ("Map_Role_Permission", len(role_perm_rows)),
            ]
        ),
        encoding="utf-8",
    )
    print(f"Generated: {OUTPUT_PATH}")
    print(f"Generated: {VERIFY_SQL_PATH}")
    print(f"Vendor source rows: {len(vendor_rows_raw)}")
    print(f"Vendor unique rows used: {len(vendor_rows)}")
    print(f"Brand rows: {len(brand_rows)}")
    print(f"Category rows: {len(category_rows)}")
    print(f"User rows: {len(user_rows)}")
    print(f"Role rows: {len(role_rows)}")
    print(f"User-role rows: {len(user_role_rows)}")
    print(f"Role-permission rows: {len(role_perm_rows)}")
    if duplicate_vendor_keys:
        print("Collapsed duplicate vendor keys:")
        for vendor_code, segment_code, payment_term_code in OrderedDict.fromkeys(duplicate_vendor_keys).keys():
            print(
                f"  VendorCode={vendor_code}, SegmentCode={segment_code}, PaymentTermCode={payment_term_code}"
            )


if __name__ == "__main__":
    main()
