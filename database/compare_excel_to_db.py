from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pyodbc
from openpyxl import load_workbook


BASE_DIR = Path(r"D:\202603_Deaktop\BMS")

EXCEL_MAPPINGS = [
    ("1.Master Vendor_BMS_v2.xlsx", "vendor", "MS_Vendor"),
    ("2.Master Brand_BMS_V1.xlsx", "brand", "MS_Brand"),
    ("3.Master category_BMS_24.11.25.xlsx", "Sheet1", "MS_Category"),
    ("Data_KBMS_User_Role.xlsx", "MS_User", "MS_User"),
    ("Data_KBMS_User_Role.xlsx", "MS_Role", "MS_Role"),
    ("Data_KBMS_User_Role.xlsx", "Map_User_Role", "Map_User_Role"),
    ("Data_KBMS_User_Role.xlsx", " Map_Role_Permission", "Map_Role_Permission"),
]


def count_excel_rows(path: Path, sheet_name: str) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    count = 0
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if row is None:
            continue
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        count += 1
    return count


def get_connection(args: argparse.Namespace) -> pyodbc.Connection:
    if args.connection_string:
        return pyodbc.connect(args.connection_string)

    if not args.server:
        raise ValueError("server is required when connection_string is not provided")

    database = args.database or "BMS"
    if args.username:
        if args.password is None:
            raise ValueError("password is required when username is provided")
        conn_str = (
            f"DRIVER={{{args.driver}}};"
            f"SERVER={args.server};DATABASE={database};UID={args.username};PWD={args.password};"
            "TrustServerCertificate=yes;"
        )
    else:
        conn_str = (
            f"DRIVER={{{args.driver}}};"
            f"SERVER={args.server};DATABASE={database};Trusted_Connection=yes;"
            "TrustServerCertificate=yes;"
        )
    return pyodbc.connect(conn_str)


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare Excel row counts with BMS SQL Server tables.")
    parser.add_argument("--connection-string", help="Full pyodbc connection string")
    parser.add_argument("--server", help="SQL Server name or host\\instance")
    parser.add_argument("--database", default="BMS", help="Database name, default BMS")
    parser.add_argument("--username", help="SQL login username")
    parser.add_argument("--password", help="SQL login password")
    parser.add_argument("--driver", default="ODBC Driver 17 for SQL Server", help="ODBC driver name")
    args = parser.parse_args()

    expected_counts: list[tuple[str, int, str, str]] = []
    for file_name, sheet_name, table_name in EXCEL_MAPPINGS:
        path = BASE_DIR / file_name
        expected = count_excel_rows(path, sheet_name)
        expected_counts.append((table_name, expected, file_name, sheet_name))

    conn = get_connection(args)
    cur = conn.cursor()

    has_fail = False
    print(f"{'Table':<22} {'Excel':>8} {'DB':>8}  Status  Source")
    print("-" * 80)
    for table_name, expected, file_name, sheet_name in expected_counts:
        row = cur.execute(f"SELECT COUNT(1) FROM dbo.[{table_name}]").fetchone()
        actual = int(row[0])
        status = "PASS" if expected == actual else "FAIL"
        if status == "FAIL":
            has_fail = True
        print(f"{table_name:<22} {expected:>8} {actual:>8}  {status:<5}  {file_name} :: {sheet_name}")

    conn.close()

    print("-" * 80)
    print("Overall:", "FAIL" if has_fail else "PASS")
    return 1 if has_fail else 0


if __name__ == "__main__":
    sys.exit(main())
